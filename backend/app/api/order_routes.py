"""
Advanced Order Mode API routes.

Endpoints:
  GET  /api/order/import          – parse & import MKCP XML files into DB
  GET  /api/order/groups          – list all vendor/stock groups
  GET  /api/order/items           – items with stock, pkg factor, reorder suggestion
  GET  /api/order/items/{name}/history  – monthly history for one item
  POST /api/order/export          – generate OrderList.xlsx download
  GET  /api/order/compliance      – basic GST compliance check on vouchers
"""
from __future__ import annotations

import io
import math
from datetime import date, datetime, timedelta
from typing import Any, Optional

from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from loguru import logger
from pydantic import BaseModel
from sqlmodel import Session, col, func, select

from app.core.config import settings
from app.core.database import get_session
from app.etl.mkcp_importer import import_mkcp
from app.models.master import StockItem
from app.models.order import AlternateUnit, ItemGroupMapping, VendorGroup
from app.models.transaction import Voucher, VoucherLine
from app.api.master_override_routes import get_all_overrides

order_router = APIRouter(prefix="/api/order", tags=["order"])

# ── Pydantic schemas ──────────────────────────────────────────────────────────


class VendorGroupRead(BaseModel):
    name: str
    parent: Optional[str]
    base_unit: str


class OrderItemRead(BaseModel):
    name: str
    group: Optional[str]
    base_unit: str
    pkg_factor: Optional[float]   # None if no price-list entry
    current_closing_base: float
    current_closing_pkg: Optional[float]
    suggestion_pkg: Optional[float]
    suggestion_base: float
    avg_monthly_outward: float


class OrderMonthlyRow(BaseModel):
    month: str          # "YYYY-MM"
    opening: float
    inward: float
    outward: float
    closing: float


class OrderExportRow(BaseModel):
    item_name: str
    group: Optional[str] = None
    qty_pkg: float = 0.0
    qty_base: float = 0.0
    uom: str = "PCS"
    current_stock: float = 0.0
    suggestion_pkg: Optional[float] = None
    remarks: str = ""


class ComplianceIssue(BaseModel):
    voucher_id: int
    voucher_number: str
    voucher_type: str
    voucher_date: str
    party_name: Optional[str]
    issues: list[str]


# ── Helpers ───────────────────────────────────────────────────────────────────


def _get_mkcp_dir() -> str:
    return getattr(settings, "MKCP_DATA_DIR", r"C:/Users/kanis/Desktop/MKCP")


def _compute_closing_maps(session: Session) -> tuple[dict, dict, dict]:
    """
    Returns three maps over ALL voucher history:
      inward_map:  {item_name → total_purchase_qty}
      outward_map: {item_name → total_sales_qty}
      unit_map:    {item_name → unit_str}
    """
    # Total purchases per item
    inward_stmt = (
        select(VoucherLine.stock_item_name, func.sum(VoucherLine.quantity), func.max(VoucherLine.unit))
        .join(Voucher, VoucherLine.voucher_id == Voucher.id)
        .where(
            func.upper(Voucher.voucher_type) == "PURCHASE",
            Voucher.is_cancelled == False,  # noqa: E712
            VoucherLine.stock_item_name.isnot(None),
        )
        .group_by(VoucherLine.stock_item_name)
    )
    inward_map: dict[str, float] = {}
    unit_map: dict[str, str] = {}
    for name, qty, unit in session.exec(inward_stmt).all():
        if name:
            inward_map[name] = float(qty or 0)
            if unit:
                unit_map[name] = unit

    # Total sales per item
    outward_stmt = (
        select(VoucherLine.stock_item_name, func.sum(VoucherLine.quantity))
        .join(Voucher, VoucherLine.voucher_id == Voucher.id)
        .where(
            func.upper(Voucher.voucher_type) == "SALES",
            Voucher.is_cancelled == False,  # noqa: E712
            VoucherLine.stock_item_name.isnot(None),
        )
        .group_by(VoucherLine.stock_item_name)
    )
    outward_map: dict[str, float] = {
        name: float(qty or 0)
        for name, qty in session.exec(outward_stmt).all()
        if name
    }

    return inward_map, outward_map, unit_map


def _compute_avg_outward_map(session: Session, lookback_months: int = 6) -> dict[str, float]:
    """Avg monthly outward (sales qty) per item over last N months."""
    cutoff = datetime.utcnow().date() - relativedelta(months=lookback_months)
    stmt = (
        select(VoucherLine.stock_item_name, func.sum(VoucherLine.quantity))
        .join(Voucher, VoucherLine.voucher_id == Voucher.id)
        .where(
            func.upper(Voucher.voucher_type) == "SALES",
            Voucher.is_cancelled == False,  # noqa: E712
            Voucher.voucher_date >= cutoff,
            VoucherLine.stock_item_name.isnot(None),
        )
        .group_by(VoucherLine.stock_item_name)
    )
    return {
        name: float(qty or 0) / max(lookback_months, 1)
        for name, qty in session.exec(stmt).all()
        if name
    }


# ── Routes ────────────────────────────────────────────────────────────────────


@order_router.get("/import")
def trigger_mkcp_import(session: Session = Depends(get_session)) -> dict:
    """Parse the 5 MKCP XML files and upsert into DB tables."""
    data_dir = _get_mkcp_dir()
    try:
        counts = import_mkcp(data_dir, session)
        logger.info(f"MKCP import complete: {counts}")
        return {"status": "ok", "data_dir": data_dir, "counts": counts}
    except Exception as exc:
        logger.error(f"MKCP import failed: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@order_router.get("/groups", response_model=list[VendorGroupRead])
def list_groups(session: Session = Depends(get_session)):
    """Return all vendor/stock groups."""
    rows = session.exec(select(VendorGroup).order_by(VendorGroup.name)).all()
    return [VendorGroupRead(name=r.name, parent=r.parent, base_unit=r.base_unit) for r in rows]


@order_router.get("/items", response_model=list[OrderItemRead])
def list_order_items(
    months_cover: int = Query(default=2, ge=1, le=12, description="Months of cover for reorder suggestion"),
    lookback: int = Query(default=6, ge=1, le=24, description="Months to average outward for suggestion"),
    group: Optional[str] = Query(default=None, description="Filter by vendor group name"),
    session: Session = Depends(get_session),
):
    """
    Return all stock items enriched with:
    - vendor group
    - alternate unit / pkg_factor
    - current closing stock (opening + purchases - sales)
    - avg monthly outward (last N months)
    - reorder suggestion (in PKG and base units)
    """
    # Build lookup dicts
    inward_map, outward_map, unit_map = _compute_closing_maps(session)
    avg_map = _compute_avg_outward_map(session, lookback)

    # Opening balances
    opening_map: dict[str, float] = {
        si.name: si.opening_balance
        for si in session.exec(select(StockItem)).all()
    }

    # AlternateUnit factors
    au_map: dict[str, AlternateUnit] = {
        au.item_name: au
        for au in session.exec(select(AlternateUnit)).all()
    }

    # Group mappings
    gm_map: dict[str, str] = {
        igm.item_name: igm.group_name
        for igm in session.exec(select(ItemGroupMapping)).all()
        if igm.group_name
    }

    # All item names (union of StockItem master + VoucherLine names)
    names_from_master = {si.name for si in session.exec(select(StockItem)).all()}
    names_from_lines = {
        r
        for r in session.exec(
            select(VoucherLine.stock_item_name)
            .distinct()
            .where(VoucherLine.stock_item_name.isnot(None))
        ).all()
    }
    all_names = sorted(names_from_master | names_from_lines)

    # Master overrides (user-editable, always take precedence)
    master_overrides = get_all_overrides()

    results: list[OrderItemRead] = []
    for name in all_names:
        item_group = gm_map.get(name)
        item_ov = master_overrides.get(name, {})

        # Apply group override
        if item_ov.get("group"):
            item_group = item_ov["group"]

        if group and item_group != group and (group != "Togo Cycles" or item_group is not None):
            continue

        opening = opening_map.get(name, 0.0)
        closing_base = opening + inward_map.get(name, 0.0) - outward_map.get(name, 0.0)

        au = au_map.get(name)
        factor: Optional[float] = au.pkg_factor if au else None

        # Apply pkg_factor override
        if item_ov.get("pkg_factor") is not None:
            factor = item_ov["pkg_factor"]

        closing_pkg: Optional[float] = None
        suggestion_pkg: Optional[float] = None
        if factor and factor > 0:
            closing_pkg = round(closing_base / factor, 2)

        avg_outward = avg_map.get(name, 0.0)
        target_base = avg_outward * months_cover
        sugg_base = max(0.0, target_base - closing_base)

        if factor and factor > 0:
            suggestion_pkg = math.ceil(sugg_base / factor)

        # Apply base_unit override
        base_unit = item_ov.get("base_unit") or unit_map.get(name, "PCS")

        results.append(
            OrderItemRead(
                name=name,
                group=item_group or "Togo Cycles",
                base_unit=base_unit,
                pkg_factor=factor,
                current_closing_base=round(closing_base, 3),
                current_closing_pkg=round(closing_pkg, 2) if closing_pkg is not None else None,
                suggestion_pkg=suggestion_pkg,
                suggestion_base=round(sugg_base, 3),
                avg_monthly_outward=round(avg_outward, 3),
            )
        )

    return results


@order_router.get("/items/{item_name}/history", response_model=list[OrderMonthlyRow])
def item_history(
    item_name: str,
    months: int = Query(default=12, ge=1, le=24),
    session: Session = Depends(get_session),
):
    """Monthly Opening / Inward / Outward / Closing for a single item."""
    si = session.exec(select(StockItem).where(StockItem.name == item_name)).first()
    opening_balance = si.opening_balance if si else 0.0

    today = datetime.utcnow().date()
    start_date = (today - relativedelta(months=months)).replace(day=1)
    current_month = start_date
    running = opening_balance
    rows: list[OrderMonthlyRow] = []

    while current_month <= today:
        month_end = (current_month + relativedelta(months=1)) - timedelta(days=1)

        inward_row = session.exec(
            select(func.sum(VoucherLine.quantity))
            .join(Voucher, VoucherLine.voucher_id == Voucher.id)
            .where(
                VoucherLine.stock_item_name == item_name,
                func.upper(Voucher.voucher_type) == "PURCHASE",
                Voucher.is_cancelled == False,  # noqa: E712
                Voucher.voucher_date >= current_month,
                Voucher.voucher_date <= month_end,
            )
        ).first()
        inward = float(inward_row or 0.0)

        outward_row = session.exec(
            select(func.sum(VoucherLine.quantity))
            .join(Voucher, VoucherLine.voucher_id == Voucher.id)
            .where(
                VoucherLine.stock_item_name == item_name,
                func.upper(Voucher.voucher_type) == "SALES",
                Voucher.is_cancelled == False,  # noqa: E712
                Voucher.voucher_date >= current_month,
                Voucher.voucher_date <= month_end,
            )
        ).first()
        outward = float(outward_row or 0.0)

        opening_this = running
        running = running + inward - outward

        rows.append(
            OrderMonthlyRow(
                month=current_month.strftime("%Y-%m"),
                opening=round(opening_this, 3),
                inward=round(inward, 3),
                outward=round(outward, 3),
                closing=round(running, 3),
            )
        )
        current_month += relativedelta(months=1)

    return rows


@order_router.post("/export")
def export_order_excel(
    rows: list[OrderExportRow],
    session: Session = Depends(get_session),
):
    """
    Generate an OrderList.xlsx from the provided order rows.
    Returns the file as a streaming download.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Order List ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "OrderList"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    row_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Item Name", "Group / Vendor", "Order Qty (PKG)", "Order Qty (PCS)",
        "UoM", "Current Stock", "Suggested (PKG)", "Remarks",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row_idx, r in enumerate(rows, 2):
        data = [
            r.item_name,
            r.group or "",
            round(r.qty_pkg, 2),
            round(r.qty_base, 3),
            r.uom,
            round(r.current_stock, 3),
            round(r.suggestion_pkg, 2) if r.suggestion_pkg is not None else "",
            r.remarks,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = row_font
            if r.qty_pkg > 0:
                cell.fill = highlight_fill

    # Auto column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ── Sheet 2: By Vendor ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Vendor")
    ws2.cell(row=1, column=1, value="Group / Vendor").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Item Name").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Order Qty (PKG)").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    # Group and sort
    from itertools import groupby

    sorted_rows = sorted(rows, key=lambda r: (r.group or "~", r.item_name))
    row_idx = 2
    for group_name, group_rows in groupby(sorted_rows, key=lambda r: r.group or "Togo Cycles"):
        group_list = list(group_rows)
        for r in group_list:
            if r.qty_pkg > 0:
                ws2.cell(row=row_idx, column=1, value=group_name)
                ws2.cell(row=row_idx, column=2, value=r.item_name)
                ws2.cell(row=row_idx, column=3, value=round(r.qty_pkg, 2))
                row_idx += 1

    for col_idx, width in [(1, 25), (2, 45), (3, 15)]:
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # Stream to client
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=OrderList.xlsx"},
    )


@order_router.get("/compliance", response_model=list[ComplianceIssue])
def gst_compliance_check(
    limit: int = Query(default=200, ge=1, le=2000),
    voucher_type: Optional[str] = Query(default=None),
    session: Session = Depends(get_session),
):
    """
    Basic GST compliance check on Sales/Purchase vouchers.
    Flags: missing invoice number, missing party, zero amount, missing HSN.
    """
    stmt = select(Voucher).where(Voucher.is_cancelled == False)  # noqa: E712
    if voucher_type:
        stmt = stmt.where(func.upper(Voucher.voucher_type) == voucher_type.upper())
    else:
        stmt = stmt.where(
            func.upper(Voucher.voucher_type).in_(["SALES", "PURCHASE"])
        )
    stmt = stmt.order_by(col(Voucher.voucher_date).desc()).limit(limit)

    issues: list[ComplianceIssue] = []
    for v in session.exec(stmt).all():
        v_issues: list[str] = []

        if not (v.voucher_number or "").strip():
            v_issues.append("Missing invoice number")
        if not (v.party_name or "").strip():
            v_issues.append("Missing party name")
        if not v.amount or v.amount <= 0:
            v_issues.append("Zero or negative amount")

        # Check HSN on at least one item line
        lines = session.exec(
            select(VoucherLine)
            .where(
                VoucherLine.voucher_id == v.id,
                VoucherLine.stock_item_name.isnot(None),
            )
            .limit(5)
        ).all()
        if lines:
            # Lookup HSN from StockItem master
            has_hsn = False
            for line in lines:
                si = session.exec(
                    select(StockItem).where(StockItem.name == line.stock_item_name)
                ).first()
                if si and si.hsn_code:
                    has_hsn = True
                    break
            if not has_hsn:
                v_issues.append("No HSN code on inventory lines")

        if v_issues:
            issues.append(
                ComplianceIssue(
                    voucher_id=v.id,
                    voucher_number=v.voucher_number or "",
                    voucher_type=v.voucher_type,
                    voucher_date=str(v.voucher_date),
                    party_name=v.party_name,
                    issues=v_issues,
                )
            )

    return issues
