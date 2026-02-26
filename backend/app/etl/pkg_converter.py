"""
PKG CONVERSION.xlsx parser — PRIMARY source for item pkg_factor.

Layout (Sheet: "Price List"):
  Row 1 : header  (S.No. | Particulars | conversion | Disc.%)
  Row 2 : item-name row   → col A = S.No. (number), col B = item name
  Row 3 : conversion row  → col A = None,            col C = pkg_factor
  Row 4 : next item-name row ...
  ...

The S.No. in col A is the reliable discriminator:
  - Non-empty number  → item-name row   (read col B as the item name)
  - Empty / None      → conversion row  (read col C as the factor)

This file is the MOST AUTHORITATIVE source for pkg_factor (items per package).
It supersedes both the Kona RATE in PRICE LIST ST.xml and the parenthetical
suffix in item names (e.g. "BELL CROWN MINI ( 300 PCS )").

Returns
-------
dict[str, float]
    Mapping of clean item name → pkg_factor (items per package).
    Names are stored as-is (stripped, original case) — matching against
    the DB uses case-insensitive lookup in the importer.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

from loguru import logger

# Normalise whitespace inside item names
_WS_RE = re.compile(r"\s+")

# Strip trailing parenthetical suffix like "( 300 PCS )", "( 1 PCS )", "( 36 )"
# These appear in xlsx names but NOT in Master.xml DB names
_PKG_SUFFIX_RE = re.compile(
    r"\s*\(\s*\d+(?:\.\d+)?\s*(?:PCS|PKG|PC|NOS|SET|PAIR|ROLL|MTR|KG|BOX)?\s*\)\s*$",
    re.IGNORECASE,
)


def _clean_name(raw: str) -> str:
    """
    Normalise an item name for DB matching:
      1. Strip trailing parenthetical suffix:  "ITEM ( 300 PCS )" → "ITEM"
      2. Collapse internal whitespace
      3. Upper-case
    """
    name = _PKG_SUFFIX_RE.sub("", raw.strip())
    return _WS_RE.sub(" ", name.strip()).upper()


def parse_pkg_conversion_xlsx(xlsx_path: str | Path) -> dict[str, float]:
    """
    Parse PKG CONVERSION.xlsx and return {clean_item_name → pkg_factor}.

    Raises
    ------
    FileNotFoundError
        If the xlsx file does not exist.
    ValueError
        If openpyxl is not installed or the expected sheet is missing.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"PKG CONVERSION xlsx not found: {xlsx_path}")

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ValueError("openpyxl is required to read .xlsx files") from exc

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # Locate sheet — prefer "Price List", fall back to first sheet
    if "Price List" in wb.sheetnames:
        ws = wb["Price List"]
    elif wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        logger.warning(
            f"pkg_converter: 'Price List' sheet not found in {xlsx_path.name}; "
            f"using first sheet '{ws.title}'"
        )
    else:
        wb.close()
        raise ValueError(f"No sheets found in {xlsx_path.name}")

    result: dict[str, float] = {}
    pending_name: Optional[str] = None
    rows_read = 0
    items_found = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        col_a = row[0]   # S.No.      — numeric on name rows, None on conversion rows
        col_b = row[1]   # Particulars / item name
        col_c = row[2]   # conversion factor (pkg_factor)
        rows_read += 1

        # ── Item-name row: S.No. is a positive number ─────────────────────────
        is_name_row = False
        if col_a is not None:
            try:
                sno = float(str(col_a).replace(",", "").strip())
                if sno > 0:
                    is_name_row = True
            except (ValueError, TypeError):
                pass

        if is_name_row:
            raw_name = str(col_b).strip() if col_b is not None else ""
            if raw_name and raw_name.lower() not in ("particulars", "none", ""):
                pending_name = _clean_name(raw_name)
                items_found += 1
            else:
                pending_name = None
            continue

        # ── Conversion row: col_c holds the factor ────────────────────────────
        if pending_name and col_c is not None:
            try:
                factor = float(str(col_c).replace(",", "").strip())
                if factor > 0:
                    result[pending_name] = factor
                else:
                    logger.debug(
                        f"pkg_converter: skipping zero/negative factor {factor!r} "
                        f"for '{pending_name}'"
                    )
            except (ValueError, TypeError):
                logger.debug(
                    f"pkg_converter: could not parse factor {col_c!r} "
                    f"for '{pending_name}'"
                )
            # Always consume pending_name (even if factor invalid)
            pending_name = None

    wb.close()

    logger.info(
        f"pkg_converter: parsed {xlsx_path.name} — "
        f"{items_found} item rows, {len(result)} factors resolved "
        f"({rows_read} total rows)"
    )
    return result


def load_pkg_conversion(data_dir: str | Path) -> dict[str, float]:
    """
    Convenience wrapper: locate PKG CONVERSION.xlsx in data_dir and parse it.

    Returns an empty dict (with a warning) if the file is not found, so the
    caller can fall back gracefully to other sources.
    """
    data_dir = Path(data_dir)
    # Try common spellings / casing
    candidates = [
        "PKG CONVERSION.xlsx",
        "pkg conversion.xlsx",
        "PKG_CONVERSION.xlsx",
        "PKG CONVERSION.xls",
    ]
    for name in candidates:
        path = data_dir / name
        if path.exists():
            try:
                return parse_pkg_conversion_xlsx(path)
            except Exception as exc:
                logger.error(f"pkg_converter: failed to parse {path.name}: {exc}")
                return {}

    logger.warning(
        f"pkg_converter: PKG CONVERSION.xlsx not found in {data_dir} "
        f"— pkg_factor will fall back to PRICE LIST ST.xml and item-name parenthetical"
    )
    return {}
